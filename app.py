import requests
import pandas as pd
import time
from openpyxl import load_workbook
from fpdf import FPDF

# API Endpoint (Using CoinGecko API as an example)
API_URL = "https://api.coingecko.com/api/v3/coins/markets"
PARAMS = {
    "vs_currency": "usd",
    "order": "market_cap_desc",
    "per_page": 50,
    "page": 1,
    "sparkline": False
}

EXCEL_FILE = "A:/deejay/crypto_data.xlsx"
REPORT_FILE = "crypto_report.pdf"
UPDATE_INTERVAL = 300  # 5 minutes

def fetch_crypto_data():
    """Fetches the top 50 cryptocurrencies by market cap."""
    response = requests.get(API_URL, params=PARAMS)
    if response.status_code == 200:
        return response.json()
    else:
        print("Error fetching data:", response.status_code)
        return None

def analyze_data(data):
    """Performs the required analysis on the cryptocurrency data."""
    df = pd.DataFrame(data)
    
    top_5 = df.nlargest(5, 'market_cap')[['name', 'market_cap']]
    avg_price = df['current_price'].mean()
    highest_change = df.loc[df['price_change_percentage_24h'].idxmax(), ['name', 'price_change_percentage_24h']]
    lowest_change = df.loc[df['price_change_percentage_24h'].idxmin(), ['name', 'price_change_percentage_24h']]
    print(top_5, avg_price, highest_change, lowest_change)
    
    return top_5, avg_price, highest_change, lowest_change

def update_excel(data):
    """Updates an Excel sheet with the live cryptocurrency data."""
    df = pd.DataFrame(data)[['name', 'symbol', 'current_price', 'market_cap', 'total_volume', 'price_change_percentage_24h']]
    df.columns = ['Name', 'Symbol', 'Current Price (USD)', 'Market Cap', '24h Volume', '24h Price Change (%)']
    
    with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl', mode='w') as writer:
        df.to_excel(writer, index=False, sheet_name="Live Data")
    
    print(f"Excel file '{EXCEL_FILE}' updated successfully.")

def generate_pdf_report(top_5, avg_price, highest_change, lowest_change):
    """Generates a PDF report summarizing key insights."""
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Arial", style="B", size=16)
    pdf.cell(200, 10, "Cryptocurrency Market Analysis Report", ln=True, align='C')
    pdf.ln(10)
    
    pdf.set_font("Arial", size=12)
    pdf.cell(200, 10, "Top 5 Cryptocurrencies by Market Cap:", ln=True)
    pdf.ln(5)
    
    for i, row in top_5.iterrows():
        pdf.cell(200, 10, f"{row['name']}: ${row['market_cap']:,}", ln=True)
    
    pdf.ln(10)
    pdf.cell(200, 10, f"Average Price of Top 50 Cryptocurrencies: ${avg_price:.2f}", ln=True)
    pdf.ln(10)
    
    pdf.cell(200, 10, f"Highest 24h Price Change: {highest_change['name']} ({highest_change['price_change_percentage_24h']:.2f}%)", ln=True)
    pdf.cell(200, 10, f"Lowest 24h Price Change: {lowest_change['name']} ({lowest_change['price_change_percentage_24h']:.2f}%)", ln=True)
    
    pdf.output(REPORT_FILE)
    print(f"PDF report '{REPORT_FILE}' generated successfully.")

def main():
    print("Fetching live cryptocurrency data...")
    data = fetch_crypto_data()
    if data:
        top_5, avg_price, highest_change, lowest_change = analyze_data(data)
        update_excel(data)
        generate_pdf_report(top_5, avg_price, highest_change, lowest_change)
    print("Process completed.")

if __name__ == "__main__":
    main()